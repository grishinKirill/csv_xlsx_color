from config import *
import openpyxl
from openpyxl import Workbook, load_workbook
#from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import re #регулярки для поиска кусков строки в ячейке
import csv #будем конвертировать csv в xlsx
import os #чисто проверим нужные библиотеки перед запуском

csv_file = 'main.csv'
if os.path.isfile(csv_file):
	wb = openpyxl.Workbook()
	ws = wb.active
else:
	print('File "%s" not found' % csv_file)
	quit()

#Конвертируем csv в xlsx
with open(csv_file) as f:
    reader = csv.reader(f, delimiter=';')
    for row in reader:
        ws.append(row)
wb_name='main.xlsx'
wb.save(wb_name) # По крайней мере файлик надо сохранить
				 # а как уже дальше пойдет- посмотрим
wb = load_workbook(wb_name)
ws = wb.active
col_len=len(ws['A:A'])
color_good="0099CC00"
color_bad="FF0000"
color_not_good_not_bad="00FF9900"

#	D:D 	FIX STATUS GREEN-RED
col_range = ws['D:D']
for i in range(1,col_len):
	val=float(re.search(r'\d{1,}',col_range[i].value)[0])
	if val == val_fix_good: col_range[i].font = Font(color=color_good, bold="True")
	else : col_range[i].font = Font(color=color_bad)

#	E:E 	FLOAT STATUS GREEN-RED
col_range = ws['E:E']
for i in range(1,col_len):
	val=float(re.search(r'\d{1,}',col_range[i].value)[0])
	if val == val_float_good: col_range[i].font = Font(color=color_good, bold="True")
	else : col_range[i].font = Font(color=color_bad)

#	F:F 	GPS STATUS GREEN-RED
col_range = ws['F:F']
for i in range(1,col_len):
	val=float(re.search(r'\d{1,}',col_range[i].value)[0])
	if val == val_gps_good: col_range[i].font = Font(color=color_good, bold="True")
	else : col_range[i].font = Font(color=color_bad)

#	G:G 	DGPS STATUS GREEN-RED
col_range = ws['G:G']
for i in range(1,col_len):
	val=float(re.search(r'\d{1,}',col_range[i].value)[0])
	if val == val_gps_good: col_range[i].font = Font(color=color_good, bold="True")
	else : col_range[i].font = Font(color=color_bad)


#	H:H 	RTCM
col_range = ws['H:H']
for i in range(1,col_len):
	val = float(re.search(r'\S{1,}\s',col_range[i].value)[0])
	if val < val_rtcm_good: col_range[i].font = Font(color=color_good)
	elif val >=val_rtcm_good and val < val_rtcm_bad : col_range[i].font = Font(color=color_not_good_not_bad)
	else: col_range[i].font = Font(color=color_bad)

#	H:H 	VOLTAGE
col_range = ws['K:K']
for i in range(1,col_len):
	volt=col_range[i].value
	val=float(volt.split('-')[1][0:-1])
	if val < val_voltage_bad: col_range[i].font = Font(color=color_bad)
	elif val >=val_voltage_bad and val < val_voltage_good: col_range[i].font = Font(color=color_not_good_not_bad)
	else: col_range[i].font = Font(color=color_good, bold="True")

#	L:L 	RSSI
col_range = ws['L:L']
for i in range(1,col_len):
	val=float(col_range[i].value.split(' ')[1])
	if val >= val_rssi_good : col_range[i].font = Font(color=color_good, bold="True")
	elif val >=val_rssi_bad and val < val_rssi_good: col_range[i].font = Font(color=color_not_good_not_bad)
	else: col_range[i].font = Font(color=color_bad)

def check_accuracy(array):
	for i in range(1,col_len):
		val_average = float(array[i].value.split(' ')[2])
		val_max = float(re.search(r'\d{1,}.\d{1,}\]{1}',array[i].value)[0][0:-1])
		if val_average < val_accuracy_av_good and val_max < val_accuracy_max_good: array[i].font = Font(color=color_good, bold="True")
		#elif val >=-80 and val < -70: col_range[i].font = Font(color=color_not_good_not_bad)
		else: array[i].font = Font(color=color_bad)
	
#	for both h_accuracy and v_accuracy
#	requirements are the same and listed 
#	in check_accuracy function

#	Q:Q 	V_ACCUR
check_accuracy(ws['Q:Q'])
#	R:R 	H_ACCUR
check_accuracy(ws['R:R'])

#print(os.time()) было бы здорово доделать уникальные имена
wb.save(wb_name)
